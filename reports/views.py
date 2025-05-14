from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.utils.decorators import method_decorator
from django.db import transaction # Import transaction

# Updated model imports
from finances.models import Income, Expense
from school.models import SchoolClass, Student, Attendance
from members.models import Member
from core.models import ChurchConfiguration # Import ChurchConfiguration
from django.utils import timezone
from datetime import date, timedelta # Added timedelta
from decimal import Decimal # Added Decimal
from dateutil.relativedelta import relativedelta
from django.db.models import Sum, Count
from django.db.models.functions import ExtractDay, TruncYear

# Import new models and forms for Accountability
from .models import AccountabilityReport, AccountabilityDocument
from .forms import AccountabilityReportForm, AccountabilityDocumentFormSet

# Imports for Export
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage # For adding images to Excel
import io
# from django.template.loader import render_to_string # No longer needed for FPDF2 for this function
from fpdf import FPDF, XPos, YPos
from django.utils.dateformat import DateFormat
from django.utils.formats import number_format
from django.conf import settings # For MEDIA_ROOT
import os # For path joining

# Helper class for FPDF reports (can be expanded)
class PDFReport(FPDF):
    def __init__(self, orientation="P", unit="mm", format="A4", church_config=None, report_title="", report_period=""):
        super().__init__(orientation, unit, format)
        self.church_config = church_config
        self.report_title = report_title
        self.report_period = report_period
        self.set_auto_page_break(True, margin=15)
        # Ensure the font path is correct for your environment (Vercel might need specific handling or font availability)
        # Using a common font like DejaVuSans is a good practice if available.
        # If /usr/share/fonts/truetype/dejavu/DejaVuSans.ttf is not available in Vercel, this will fail.
        # Consider bundling the font or using a path that Verce        
        try:
        # Attempt to add and use DejaVu font
            self.add_font("DejaVu", "", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", uni=True)
            self.add_font("DejaVu", "B", "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", uni=True)
            self.set_font("DejaVu", "", 10)
        except (RuntimeError, FileNotFoundError) as e: # Catch both RuntimeError and FileNotFoundError
            print(f"FPDF Font Error (DejaVu not found or load failed): {e}. Falling back to Helvetica.")
            # Fallback to a standard FPDF font if DejaVu is not available or fails to load
            self.set_font("Helvetica", "", 10)
            # If you need a bold version for Helvetica, you can set it when needed, e.g., self.set_font("Helvetica", "B", 10)
            # FPDF's standard fonts (Helvetica, Times, Courier) don't require add_font for bold/italic styles, just set_font.ef    

    def header(self):
        if not self.church_config:
            return

        page_width = self.w - 2 * self.l_margin
        logo_width = 20
        text_area_width = page_width - logo_width - 5 # 5mm spacing

        # Logo
        if self.church_config.logo and hasattr(self.church_config.logo, "path") and os.path.exists(self.church_config.logo.path):
            try:
                self.image(self.church_config.logo.path, self.l_margin, 10, logo_width)
            except Exception as e:
                print(f"Error adding logo to PDF header: {e}")
        
        # Church Name
        current_font_family = self.font_family
        current_font_style = self.font_style
        current_font_size = self.font_size_pt
        
        self.set_xy(self.l_margin + logo_width + 5, 10)
        self.set_font(current_font_family, 'B', 16)
        self.multi_cell(text_area_width, 8, self.church_config.church_name or "Nome da Igreja", 0, 'C')

        # Pastor and         self.set_xy(self.l_margin + logo_width + 5, self.get_y())
        self.set_font(current_font_family, '', 9)
        pastor_treasurer_text = f"Pastor: {self.church_config.president_pastor_name or '-'} | Tesoureiro(a): {self.church_config.treasurer_name or '-'}"
        self.multi_cell(text_area_width, 5, pastor_treasurer_text, 0, 'C')
        self.ln(5)

        # Report Title
        self.set_x(self.l_margin)
        self.set_font(current_font_family, 'B', 12)
        self.cell(page_width, 10, self.report_title, 0, 1, 'C')
        
        # Report Perio        self.set_x(self.l_margin)
        self.set_font(current_font_family, '', 10)
        self.cell(page_width, 7, self.report_period, 0, 1, 'C')
        self.ln(5) # Space after header
        self.set_font(current_font_family, current_font_style, current_font_size) # Reset font

    def footer(self):
        current_font_family = self.font_family
        current_font_style = self.font_style
        # current_font_size = self.font_size_pt

        self.set_y(-15)
        self.set_font(current_font_family, '', 8)
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, 'C')
        today_str = DateFormat(date.today()).format("d/m/Y")
        self.set_x(self.l_margin)
        self.cell(0,10, f"Gerado em: {today_str}", 0,0, 'L')
        # self.set_font(current_font_family, current_font_style, current_font_size) # Reset font


@login_required
def index(request):
    context = {
        "active_menu": "reports",
    }
    return render(request, "reports/index.html", context)

# --- Helper function to get common context for reports ---
def _get_report_filters(request):
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year

    try:
        month_param = int(request.GET.get("month", current_month))
        year_param = int(request.GET.get("year", current_year))
        filter_date = date(year_param, month_param, 1)
    except (ValueError, TypeError):
        month_param = current_month
        year_param = current_year
        filter_date = date(current_year, current_month, 1)

    try:
        end_date_str = request.GET.get("end_date", today.strftime("%Y-%m-%d"))
        end_date = date.fromisoformat(end_date_str)
    except (ValueError, TypeError):
        end_date = today
        end_date_str = end_date.strftime("%Y-%m-%d")

    class_id = request.GET.get("class_id")
    try:
        class_date_str = request.GET.get("class_date", today.strftime("%Y-%m-%d"))
        class_date = date.fromisoformat(class_date_str)
    except (ValueError, TypeError):
        class_date = today
        class_date_str = class_date.strftime("%Y-%m-%d")

    member_param = request.GET.get("member")
    
    church_config = ChurchConfiguration.objects.first()

    return {
        "today": today,
        "current_month": current_month,
        "current_year": current_year,
        "month_param": month_param,
        "year_param": year_param,
        "filter_date": filter_date,
        "end_date": end_date,
        "end_date_str": end_date_str,
        "class_id": class_id,
        "class_date": class_date,
        "class_date_str": class_date_str,
        "member_param": member_param,
        "church_config": church_config,
        "available_years": range(current_year, current_year - 5, -1),
        "available_months": {
            1: "Janeiro", 2: "Fevereiro", 3: "Março", 4: "Abril",
            5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
            9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro"
        }
    }


# --- Relatórios Financeiros ---
@login_required
def relatorio_movimentacoes_mensais(request):
    filters = _get_report_filters(request)
    first_day_month = filters["filter_date"]
    last_day_month = first_day_month + relativedelta(months=1) - relativedelta(days=1)
    
    incomes = Income.objects.filter(date__gte=first_day_month, date__lte=last_day_month).order_by("date")
    expenses = Expense.objects.filter(date__gte=first_day_month, date__lte=last_day_month).order_by("date")

    total_incomes = sum(i.amount for i in incomes)
    total_expenses = sum(e.amount for e in expenses)
    month_balance = total_incomes - total_expenses

    context = {
        "active_menu": "reports",
        "incomes": incomes, 
        "expenses": expenses, 
        "total_incomes": total_incomes, 
        "total_expenses": total_expenses, 
        "month_balance": month_balance, 
        "selected_month": first_day_month.month, 
        "selected_year": first_day_month.year, 
        "month_name": filters["available_months"].get(first_day_month.month),
        "available_years": filters["available_years"],
        "available_months": filters["available_months"],
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }
    return render(request, "reports/movimentacoes_mensais.html", context)


@login_required
def export_movimentacoes_mensais_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    first_day_month = filters["filter_date"]
    last_day_month = first_day_month + relativedelta(months=1) - relativedelta(days=1)

    incomes = Income.objects.filter(date__gte=first_day_month, date__lte=last_day_month).order_by("date")
    expenses = Expense.objects.filter(date__gte=first_day_month, date__lte=last_day_month).order_by("date")
    total_incomes = sum(i.amount for i in incomes)
    total_expenses = sum(e.amount for e in expenses)
    month_balance = total_incomes - total_expenses

    wb = Workbook()
    ws = wb.active
    ws.title = f"Mov. {first_day_month.strftime('%b_%Y')}"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75 
                img.width = 75  
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor Presidente: {church_config.president_pastor_name or '-'} | Tesoureiro(a): {church_config.treasurer_name or '-'}")
        cell.font = Font(size=10)
        cell.alignment = Alignment(horizontal="center")
        current_row += 1
    title_cell = ws.cell(row=current_row, column=2)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    ws.cell(row=current_row, column=1, value="Receitas").font = Font(bold=True, size=12)
    current_row += 1
    header_incomes = ["Data", "Descrição", "Categoria", "Membro", "Valor"]
    ws.append(header_incomes)
    header_font = Font(bold=True)
    for col_idx, _ in enumerate(header_incomes, 1):
        ws.cell(row=current_row, column=col_idx).font = header_font
    current_row += 1

    for income_item in incomes:
        ws.append([
            DateFormat(income_item.date).format("d/m/Y"),
            income_item.description,
            income_item.category.name if income_item.category else "",
            income_item.member.name if income_item.member else "",
            income_item.amount
        ])
        ws.cell(row=current_row, column=5).number_format = '"R$" #,##0.00'
        current_row += 1

    ws.append(["", "", "", "Total Receitas:", total_incomes])
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5).font = Font(bold=True)
    ws.cell(row=current_row, column=5).number_format = '"R$" #,##0.00'
    current_row += 2

    ws.cell(row=current_row, column=1, value="Despesas").font = Font(bold=True, size=12)
    current_row += 1
    header_expenses = ["Data", "Descrição", "Categoria", "", "Valor"]
    ws.append(header_expenses)
    for col_idx, _ in enumerate(header_expenses, 1):
        if ws.cell(row=current_row, column=col_idx).value:
             ws.cell(row=current_row, column=col_idx).font = header_font
    current_row += 1

    for expense_item in expenses:
        ws.append([
            DateFormat(expense_item.date).format("d/m/Y"),
            expense_item.description,
            expense_item.category.name if expense_item.category else "",
            "",
            expense_item.amount
        ])
        ws.cell(row=current_row, column=5).number_format = '"R$" #,##0.00'
        current_row += 1

    ws.append(["", "", "", "Total Despesas:", total_expenses])
    ws.cell(row=current_row, column=4).font = Font(bold=True)
    ws.cell(row=current_row, column=5).font = Font(bold=True)
    ws.cell(row=current_row, column=5).number_format = '"R$" #,##0.00'
    current_row += 2

    ws.append(["", "", "", "Saldo do Mês:", month_balance])
    ws.cell(row=current_row, column=4).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=5).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=5).number_format = '"R$" #,##0.00'

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45 # Increased width for description
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=movimentacoes_{first_day_month.strftime("%Y_%m")}.xlsx"
    return response

@login_required
def export_movimentacoes_mensais_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    first_day_month = filters["filter_date"]
    report_period_str = f"{filters['available_months'].get(first_day_month.month)} {first_day_month.year}"

    incomes_qs = Income.objects.filter(date__gte=first_day_month, date__lte=first_day_month + relativedelta(months=1) - relativedelta(days=1)).order_by("date")
    expenses_qs = Expense.objects.filter(date__gte=first_day_month, date__lte=first_day_month + relativedelta(months=1) - relativedelta(days=1)).order_by("date")
    total_incomes = sum(i.amount for i in incomes_qs)
    total_expenses = sum(e.amount for e in expenses_qs)
    month_balance = total_incomes - total_expenses

    pdf = PDFReport(church_config=church_config, 
                    report_title="Relatório de Movimentações Mensais", 
                    report_period=report_period_str)
    pdf.add_page()
    current_font_family = pdf.font_family # Store current font to reset after section

    col_widths = {
        "date": 25,
        "description": 70,
        "category": 35,
        "member": 35,
        "value": 25
    }
    line_height = 7
    page_width = pdf.w - 2 * pdf.l_margin

    # --- Incomes Section ---
    pdf.set_font(current_font_family, 'B', 11)
    pdf.cell(page_width, line_height * 1.5, "Receitas", 0, 1, 'L')
    pdf.set_font(current_font_family, 'B', 9)
    header_incomes = ["Data", "Descrição", "Categoria", "Membro", "Valor"]
    current_x = pdf.l_margin
    pdf.set_x(current_x)
    pdf.cell(col_widths["date"], line_height, header_incomes[0], 1, 0, 'C')
    current_x += col_widths["date"]
    pdf.set_x(current_x)
    pdf.cell(col_widths["description"], line_height, header_incomes[1], 1, 0, 'C')
    current_x += col_widths["description"]
    pdf.set_x(current_x)
    pdf.cell(col_widths["category"], line_height, header_incomes[2], 1, 0, 'C')
    current_x += col_widths["category"]
    pdf.set_x(current_x)
    pdf.cell(col_widths["member"], line_height, header_incomes[3], 1, 0, 'C')
    current_x += col_widths["member"]
    pdf.set_x(current_x)
    pdf.cell(col_widths["value"], line_height, header_incomes[4], 1, 1, 'C') # New line

    pdf.set_font(current_font_family, '', 8)
    for item in incomes_qs:
        start_y = pdf.get_y()
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(col_widths["date"], line_height, DateFormat(item.date).format("d/m/Y"), 1, 'L')
        max_y = pdf.get_y()
        pdf.set_xy(pdf.l_margin + col_widths["date"], start_y)
        
        pdf.multi_cell(col_widths["description"], line_height, item.description or "", 1, 'L')
        max_y = max(max_y, pdf.get_y())
        pdf.set_xy(pdf.l_margin + col_widths["date"] + col_widths["description"], start_y)

        pdf.multi_cell(col_widths["category"], line_height, item.category.name if item.category else "", 1, 'L')
        max_y = max(max_y, pdf.get_y())
        pdf.set_xy(pdf.l_margin + col_widths["date"] + col_widths["description"] + col_widths["category"], start_y)

        pdf.multi_cell(col_widths["member"], line_height, item.member.name if item.member else "", 1, 'L')
        max_y = max(max_y, pdf.get_y())
        pdf.set_xy(pdf.l_margin + col_widths["date"] + col_widths["description"] + col_widths["category"] + col_widths["member"], start_y)
        
        pdf.multi_cell(col_widths["value"], line_height, f"R$ {number_format(item.amount, decimal_pos=2, force_grouping=True)}", 1, 'R')
        max_y = max(max_y, pdf.get_y())
        pdf.set_y(max_y) # Move to the bottom of the tallest cell in the row

    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin + col_widths["date"] + col_widths["description"] + col_widths["category"])
    pdf.cell(col_widths["member"], line_height, "Total Receitas:", 1, 0, 'R')
    pdf.cell(col_widths["value"], line_height, f"R$ {number_format(total_incomes, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    pdf.ln(line_height / 2)

    # --- Expenses Section ---
    pdf.set_font(current_font_family, 'B', 11)
    pdf.set_x(pdf.l_margin)
    pdf.cell(page_width, line_height * 1.5, "Despesas", 0, 1, 'L')
    pdf.set_font(current_font_family, 'B', 9)
    header_expenses = ["Data", "Descrição", "Categoria", "Valor"]
    current_x = pdf.l_margin
    pdf.set_x(current_x)
    pdf.cell(col_widths["date"], line_height, header_expenses[0], 1, 0, 'C')
    current_x += col_widths["date"]
    pdf.set_x(current_x)
    expense_desc_width = col_widths["description"] + col_widths["member"] # Combine for wider description
    pdf.cell(expense_desc_width, line_height, header_expenses[1], 1, 0, 'C')
    current_x += expense_desc_width
    pdf.set_x(current_x)
    pdf.cell(col_widths["category"], line_height, header_expenses[2], 1, 0, 'C')
    current_x += col_widths["category"]
    pdf.set_x(current_x)
    pdf.cell(col_widths["value"], line_height, header_expenses[3], 1, 1, 'C') # New line

    pdf.set_font(current_font_family, '', 8)
    for item in expenses_qs:
        start_y = pdf.get_y()
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(col_widths["date"], line_height, DateFormat(item.date).format("d/m/Y"), 1, 'L')
        max_y = pdf.get_y()
        pdf.set_xy(pdf.l_margin + col_widths["date"], start_y)

        pdf.multi_cell(expense_desc_width, line_height, item.description or "", 1, 'L')
        max_y = max(max_y, pdf.get_y())
        pdf.set_xy(pdf.l_margin + col_widths["date"] + expense_desc_width, start_y)

        pdf.multi_cell(col_widths["category"], line_height, item.category.name if item.category else "", 1, 'L')
        max_y = max(max_y, pdf.get_y())
        pdf.set_xy(pdf.l_margin + col_widths["date"] + expense_desc_width + col_widths["category"], start_y)
        
        pdf.multi_cell(col_widths["value"], line_height, f"R$ {number_format(item.amount, decimal_pos=2, force_grouping=True)}", 1, 'R')
        max_y = max(max_y, pdf.get_y())
        pdf.set_y(max_y)

    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin + col_widths["date"] + expense_desc_width)
    pdf.cell(col_widths["category"], line_height, "Total Despesas:", 1, 0, 'R')
    pdf.cell(col_widths["value"], line_height, f"R$ {number_format(total_expenses, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    pdf.ln(line_height)

    # --- Balance Section ---
    pdf.set_font(current_font_family, 'B', 10)
    pdf.set_x(pdf.l_margin + col_widths["date"] + expense_desc_width)
    pdf.cell(col_widths["category"], line_height, "Saldo do Mês:", 1, 0, 'R')
    pdf.cell(col_widths["value"], line_height, f"R$ {number_format(month_balance, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')

    pdf.set_font(current_font_family, '', 10) # Reset font to default for subsequent calls if any
    pdf_output_bytes = pdf.output(dest='S')

    response = HttpResponse(pdf_output_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="movimentacoes_{first_day_month.strftime("%Y_%m")}.pdf"'
    return response


@login_required
def relatorio_dre(request):
    filters = _get_report_filters(request)
    year_param = filters["year_param"]
    first_day_year = date(year_param, 1, 1)
    last_day_year = date(year_param, 12, 31)

    incomes_by_category = Income.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")

    expenses_by_category = Expense.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")

    total_revenue = sum(item["total"] for item in incomes_by_category if item["total"]) or 0
    total_expenditure = sum(item["total"] for item in expenses_by_category if item["total"]) or 0
    net_result = total_revenue - total_expenditure

    context = {
        "active_menu": "reports",
        "selected_year": year_param,
        "incomes_by_category": incomes_by_category, 
        "expenses_by_category": expenses_by_category, 
        "total_revenue": total_revenue, 
        "total_expenditure": total_expenditure, 
        "net_result": net_result, 
        "available_years": filters["available_years"],
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }
    return render(request, "reports/dre.html", context)

@login_required
def export_dre_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    year_param = filters["year_param"]
    first_day_year = date(year_param, 1, 1)
    last_day_year = date(year_param, 12, 31)

    incomes_by_category = Income.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")
    expenses_by_category = Expense.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")
    total_revenue = sum(item["total"] for item in incomes_by_category if item["total"]) or 0
    total_expenditure = sum(item["total"] for item in expenses_by_category if item["total"]) or 0
    net_result = total_revenue - total_expenditure

    wb = Workbook()
    ws = wb.active
    ws.title = f"DRE {year_param}"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75; img.width = 75
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to DRE Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16); cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor: {church_config.president_pastor_name or "-"} | Tesoureiro: {church_config.treasurer_name or "-"}")
        cell.font = Font(size=10); cell.alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1, value=f"Demonstração do Resultado do Exercício - {year_param}")
    title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    ws.cell(row=current_row, column=1, value="Receitas Operacionais").font = Font(bold=True, size=12)
    current_row +=1
    header_dre_incomes = ["Categoria", "Total"]
    ws.append(header_dre_incomes)
    header_font = Font(bold=True)
    for col_idx, _ in enumerate(header_dre_incomes, 1):
        ws.cell(row=current_row, column=col_idx).font = header_font
    current_row +=1

    for item in incomes_by_category:
        ws.append([
            item["category__name"] or "N/A",
            item["total"]
        ])
        ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
        current_row +=1
    
    ws.append(["Total Receitas Operacionais:", total_revenue])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
    current_row += 2

    ws.cell(row=current_row, column=1, value="Despesas Operacionais").font = Font(bold=True, size=12)
    current_row +=1
    header_dre_expenses = ["Categoria", "Total"]
    ws.append(header_dre_expenses)
    for col_idx, _ in enumerate(header_dre_expenses, 1):
        ws.cell(row=current_row, column=col_idx).font = header_font
    current_row +=1

    for item in expenses_by_category:
        ws.append([
            item["category__name"] or "N/A",
            item["total"]
        ])
        ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
        current_row +=1

    ws.append(["Total Despesas Operacionais:", total_expenditure])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).font = Font(bold=True)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
    current_row += 2

    ws.append(["Resultado Líquido do Exercício:", net_result])
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=DRE_{year_param}.xlsx"
    return response

@login_required
def export_dre_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    year_param = filters["year_param"]
    report_period_str = f"Ano de {year_param}"

    first_day_year = date(year_param, 1, 1)
    last_day_year = date(year_param, 12, 31)

    incomes_by_category = Income.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")
    expenses_by_category = Expense.objects.filter(
        date__gte=first_day_year, date__lte=last_day_year
    ).values("category__name").annotate(total=Sum("amount")).order_by("-total")
    total_revenue = sum(item["total"] for item in incomes_by_category if item["total"]) or 0
    total_expenditure = sum(item["total"] for item in expenses_by_category if item["total"]) or 0
    net_result = total_revenue - total_expenditure

    pdf = PDFReport(church_config=church_config, 
                    report_title="Demonstração do Resultado do Exercício (DRE)", 
                    report_period=report_period_str)
    pdf.add_page()
    current_font_family = pdf.font_family # Store current font

    col_width_category = 140
    col_width_value = 50
    line_height = 7
    page_width = pdf.w - 2 * pdf.l_margin

    # --- Incomes Section ---
    pdf.set_font(current_font_family, 'B', 11)
    pdf.set_x(pdf.l_margin)
    pdf.cell(page_width, line_height * 1.5, "Receitas Operacionais", 0, 1, 'L')
    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_category, line_height, "Categoria", 1, 0, 'C')
    pdf.cell(col_width_value, line_height, "Total", 1, 1, 'C')

    pdf.set_font(current_font_family, '', 8)
    for item in incomes_by_category:
        start_y = pdf.get_y()
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(col_width_category, line_height, item["category__name"] or "N/A", 1, 'L')
        max_y = pdf.get_y()
        pdf.set_xy(pdf.l_margin + col_width_category, start_y)
        pdf.multi_cell(col_width_value, line_height, f"R$ {number_format(item['total'], decimal_pos=2, force_grouping=True)}", 1, 'R')
        max_y = max(max_y, pdf.get_y())
        pdf.set_y(max_y)
    
    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_category, line_height, "Total Receitas Operacionais:", 1, 0, 'R')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(total_revenue, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    pdf.ln(line_height / 2)

    # --- Expenses Section ---
    pdf.set_font(current_font_family, 'B', 11)
    pdf.set_x(pdf.l_margin)
    pdf.cell(page_width, line_height * 1.5, "Despesas Operacionais", 0, 1, 'L')
    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_category, line_height, "Categoria", 1, 0, 'C')
    pdf.cell(col_width_value, line_height, "Total", 1, 1, 'C')

    pdf.set_font(current_font_family, '', 8)
    for item in expenses_by_category:
        start_y = pdf.get_y()
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(col_width_category, line_height, item["category__name"] or "N/A", 1, 'L')
        max_y = pdf.get_y()
        pdf.set_xy(pdf.l_margin + col_width_category, start_y)
        pdf.multi_cell(col_width_value, line_height, f"R$ {number_format(item['total'], decimal_pos=2, force_grouping=True)}", 1, 'R')
        max_y = max(max_y, pdf.get_y())
        pdf.set_y(max_y)

    pdf.set_font(current_font_family, 'B', 9)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_category, line_height, "Total Despesas Operacionais:", 1, 0, 'R')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(total_expenditure, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    pdf.ln(line_height)

    # --- Net Result Section ---
    pdf.set_font(current_font_family, 'B', 10)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_category, line_height, "Resultado Líquido do Exercício:", 1, 0, 'R')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(net_result, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')

    pdf.set_font(current_font_family, '', 10) # Reset font
    pdf_output_bytes = pdf.output(dest='S')
    response = HttpResponse(pdf_output_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="DRE_{year_param}.pdf"'
    return response


@login_required
def relatorio_balanco_rever(request):
    filters = _get_report_filters(request)
    year_param = filters["year_param"]
    first_day_year = date(year_param, 1, 1)
    last_day_year = date(year_param, 12, 31)

    total_assets = Income.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    total_liabilities = Expense.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    equity = total_assets - total_liabilities # Simplified equity

    context = {
        "active_menu": "reports",
        "selected_year": year_param,
        "total_assets": total_assets,
        "total_liabilities_equity": total_liabilities,
        "equity": equity,
        "available_years": filters["available_years"],
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }
    return render(request, "reports/balanco.html", context)

@login_required
def relatorio_balanco(request):
    filters = _get_report_filters(request)
    end_date = filters["end_date"]
    total_incomes_accumulated = Income.objects.filter(date__lte=end_date).aggregate(total=Sum("amount"))["total"] or 0
    total_expenses_accumulated = Expense.objects.filter(date__lte=end_date).aggregate(total=Sum("amount"))["total"] or 0
    accumulated_balance = total_incomes_accumulated - total_expenses_accumulated
    assets = {"Caixa/Banco (Saldo Acumulado)": accumulated_balance}
    liabilities_equity = {"Patrimônio Líquido (Resultado Acumulado)": accumulated_balance}

    context = {
        "active_menu": "reports",
        "end_date": end_date, 
        "end_date_str": filters["end_date_str"],
        "assets": assets,
        "liabilities_equity": liabilities_equity, 
        "total_assets": sum(assets.values()),
        "total_liabilities_equity": sum(liabilities_equity.values()), 
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }
    return render(request, "reports/balanco.html", context)

@login_required
def export_balanco_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    year_param = filters["year_param"]
    
    total_assets = Income.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    total_liabilities = Expense.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    equity = total_assets - total_liabilities

    wb = Workbook()
    ws = wb.active
    ws.title = f"Balanco {year_param}"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75; img.width = 75
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to Balanço Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16); cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor: {church_config.president_pastor_name or "-"} | Tesoureiro: {church_config.treasurer_name or "-"}")
        cell.font = Font(size=10); cell.alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1, value=f"Balanço Patrimonial Simplificado - {year_param}")
    title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    ws.append(["Ativos (Total Receitas do Ano)", total_assets])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
    current_row +=1
    ws.append(["Passivos (Total Despesas do Ano)", total_liabilities])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'
    current_row +=1
    ws.append(["Patrimônio Líquido (Simplificado)", equity])
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=2).font = Font(bold=True, size=12)
    ws.cell(row=current_row, column=2).number_format = '"R$" #,##0.00'

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Balanco_Patrimonial_{year_param}.xlsx"
    return response

@login_required
def export_balanco_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    year_param = filters["year_param"]
    report_period_str = f"Em 31 de Dezembro de {year_param}"

    total_assets = Income.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    total_liabilities = Expense.objects.filter(date__year=year_param).aggregate(Sum('amount'))['amount__sum'] or Decimal(0)
    equity = total_assets - total_liabilities

    pdf = PDFReport(church_config=church_config, 
                    report_title="Balanço Patrimonial Simplificado", 
                    report_period=report_period_str)
    pdf.add_page()
    current_font_family = pdf.font_family

    col_width_desc = 140
    col_width_value = 50
    line_height = 8

    pdf.set_font(current_font_family, 'B', 10)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_desc, line_height, "Ativos (Total Receitas do Ano)", 1, 0, 'L')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(total_assets, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_desc, line_height, "Passivos (Total Despesas do Ano)", 1, 0, 'L')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(total_liabilities, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')
    pdf.ln(line_height / 2)

    pdf.set_font(current_font_family, 'B', 11)
    pdf.set_x(pdf.l_margin)
    pdf.cell(col_width_desc, line_height, "Patrimônio Líquido (Simplificado)", 1, 0, 'L')
    pdf.cell(col_width_value, line_height, f"R$ {number_format(equity, decimal_pos=2, force_grouping=True)}", 1, 1, 'R')

    pdf.set_font(current_font_family, '', 10)
    pdf_output_bytes = pdf.output(dest='S')
    response = HttpResponse(pdf_output_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Balanco_Patrimonial_{year_param}.pdf"'
    return response

# --- Relatórios da Escola Dominical ---
@login_required
def relatorio_alunos_por_turma(request):
    filters = _get_report_filters(request)
    classes = SchoolClass.objects.all().prefetch_related('students')
    context = {
        "active_menu": "reports",
        "classes": classes,
        "church_config": filters["church_config"],
    }
    return render(request, "reports/alunos_por_turma.html", context)

@login_required
def export_alunos_por_turma_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    classes = SchoolClass.objects.all().prefetch_related('students')

    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos por Turma"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75; img.width = 75
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to Alunos/Turma Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16); cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor: {church_config.president_pastor_name or "-"} | Tesoureiro: {church_config.treasurer_name or "-"}")
        cell.font = Font(size=10); cell.alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    title_cell = ws.cell(row=current_row, column=1, value="Relatório de Alunos por Turma")
    title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    header_font = Font(bold=True)
    for school_class in classes:
        ws.cell(row=current_row, column=1, value=f"Turma: {school_class.name}").font = Font(bold=True, size=12)
        current_row += 1
        ws.cell(row=current_row, column=1, value=f"Professor: {school_class.teacher_name}").font = Font(italic=True)
        current_row += 1
        
        headers = ["Nome do Aluno", "Data de Nascimento", "Telefone", "Email"]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_idx).font = header_font
        current_row += 1

        for student in school_class.students.all():
            ws.append([
                student.name,
                DateFormat(student.birth_date).format("d/m/Y") if student.birth_date else "",
                student.phone_number,
                student.email
            ])
            current_row +=1
        ws.append([]) # Spacer row
        current_row +=1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=alunos_por_turma.xlsx"
    return response

@login_required
def export_alunos_por_turma_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    classes = SchoolClass.objects.all().prefetch_related('students')

    pdf = PDFReport(church_config=church_config, 
                    report_title="Relatório de Alunos por Turma", 
                    report_period=f"Gerado em: {DateFormat(date.today()).format('d/m/Y')}")
    pdf.add_page()
    current_font_family = pdf.font_family

    line_height = 6
    col_widths_students = {"name": 70, "birth_date": 30, "phone": 40, "email": 50}

    for school_class in classes:
        if pdf.get_y() > (pdf.h - pdf.b_margin - 40): 
            pdf.add_page()

        pdf.set_font(current_font_family, 'B', 11)
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, line_height * 1.5, f"Turma: {school_class.name}", 0, 1, 'L')
        pdf.set_font(current_font_family, '', 9)
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, line_height, f"Professor: {school_class.teacher_name}", 0, 1, 'L')
        pdf.ln(line_height / 2)

        pdf.set_font(current_font_family, 'B', 8)
        pdf.set_x(pdf.l_margin)
        pdf.cell(col_widths_students["name"], line_height, "Nome do Aluno", 1, 0, 'C')
        pdf.cell(col_widths_students["birth_date"], line_height, "Dt. Nasc.", 1, 0, 'C')
        pdf.cell(col_widths_students["phone"], line_height, "Telefone", 1, 0, 'C')
        pdf.cell(col_widths_students["email"], line_height, "Email", 1, 1, 'C')

        pdf.set_font(current_font_family, '', 7)
        for student in school_class.students.all():
            if pdf.get_y() > (pdf.h - pdf.b_margin - line_height * 2):
                 pdf.add_page()
                 pdf.set_font(current_font_family, 'B', 8)
                 pdf.set_x(pdf.l_margin)
                 pdf.cell(col_widths_students["name"], line_height, "Nome do Aluno", 1, 0, 'C')
                 pdf.cell(col_widths_students["birth_date"], line_height, "Dt. Nasc.", 1, 0, 'C')
                 pdf.cell(col_widths_students["phone"], line_height, "Telefone", 1, 0, 'C')
                 pdf.cell(col_widths_students["email"], line_height, "Email", 1, 1, 'C')
                 pdf.set_font(current_font_family, '', 7)

            start_y = pdf.get_y()
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(col_widths_students["name"], line_height, student.name, 1, 'L')
            max_y = pdf.get_y()
            pdf.set_xy(pdf.l_margin + col_widths_students["name"], start_y)
            
            pdf.multi_cell(col_widths_students["birth_date"], line_height, DateFormat(student.birth_date).format("d/m/Y") if student.birth_date else "", 1, 'L')
            max_y = max(max_y, pdf.get_y())
            pdf.set_xy(pdf.l_margin + col_widths_students["name"] + col_widths_students["birth_date"], start_y)

            pdf.multi_cell(col_widths_students["phone"], line_height, student.phone_number or "", 1, 'L')
            max_y = max(max_y, pdf.get_y())
            pdf.set_xy(pdf.l_margin + col_widths_students["name"] + col_widths_students["birth_date"] + col_widths_students["phone"], start_y)

            pdf.multi_cell(col_widths_students["email"], line_height, student.email or "", 1, 'L')
            max_y = max(max_y, pdf.get_y())
            pdf.set_y(max_y)
        pdf.ln(line_height) 

    pdf.set_font(current_font_family, '', 10)
    pdf_output_bytes = pdf.output(dest='S')
    response = HttpResponse(pdf_output_bytes, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="alunos_por_turma.pdf"'
    return response

@login_required
def relatorio_frequencia(request):
    filters = _get_report_filters(request)
    class_id = filters["class_id"]
    class_date = filters["class_date"]
    classes = SchoolClass.objects.all().order_by("name")
    attendances = Attendance.objects.filter(date=class_date).select_related("student__member", "school_class")
    selected_class = None
    if class_id:
        try:
            selected_class = SchoolClass.objects.get(pk=class_id)
            attendances = attendances.filter(school_class_id=class_id)
        except SchoolClass.DoesNotExist:
            class_id = None # Reset if class not found

    attendances = attendances.order_by("school_class__name", "student__member__name")
    total_present = attendances.filter(present=True).count()
    total_absent = attendances.filter(present=False).count()
    total_records = attendances.count()

    context = {
        "active_menu": "reports",
        "classes": classes,
        "selected_class_id": class_id,
        "selected_class_name": selected_class.name if selected_class else "Todas as Turmas",
        "class_date": class_date,
        "class_date_str": filters["class_date_str"],
        "attendances": attendances,
        "total_present": total_present,
        "total_absent": total_absent,
        "total_records": total_records,
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }
    return render(request, "reports/frequencia.html", context)

@login_required
def export_frequencia_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    class_id = filters["class_id"]
    class_date = filters["class_date"]
    class_date_str = filters["class_date_str"]

    selected_class = None
    attendance_records = []
    if class_id:
        try:
            selected_class = SchoolClass.objects.get(pk=class_id)
            students_in_class = selected_class.students.all().order_by('name')
            recorded_attendance = Attendance.objects.filter(school_class=selected_class, date=class_date)
            present_student_ids = list(recorded_attendance.values_list('student_id', flat=True))
            for student in students_in_class:
                attendance_records.append({
                    'student_name': student.name,
                    'status': 'Presente' if student.id in present_student_ids else 'Ausente'
                })
        except SchoolClass.DoesNotExist:
            pass 

    wb = Workbook()
    ws = wb.active
    ws.title = f"Frequencia_{selected_class.name[:10] if selected_class else ''}_{class_date_str.replace('-' , '')}"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75; img.width = 75
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to Frequencia Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16); cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor: {church_config.president_pastor_name or "-"} | Tesoureiro: {church_config.treasurer_name or "-"}")
        cell.font = Font(size=10); cell.alignment = Alignment(horizontal="center")
        current_row += 1

    report_title_str = "Relatório de Frequência"
    if selected_class:
        report_title_str += f" - Turma: {selected_class.name}"
    report_title_str += f" - Data: {DateFormat(class_date).format('d/m/Y')}"
    
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
    title_cell = ws.cell(row=current_row, column=1, value=report_title_str)
    title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    if attendance_records:
        headers = ["Nome do Aluno", "Status"]
        ws.append(headers)
        header_font = Font(bold=True)
        for col_idx, _ in enumerate(headers, 1):
            ws.cell(row=current_row, column=col_idx).font = header_font
        current_row += 1
        for record in attendance_records:
            ws.append([record['student_name'], record['status']])
            current_row +=1
    else:
        ws.append(["Nenhum dado de frequência para os filtros selecionados."])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=frequencia_{selected_class.name.replace(' ', '') if selected_class else 'turma'}_{class_date_str.replace('-' , '')}.xlsx"
    return response

@login_required
def export_frequencia_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    class_id = filters["class_id"]
    class_date = filters["class_date"]
    
    selected_class = None
    attendance_records = []
    if class_id:
        try:
            selected_class = SchoolClass.objects.get(pk=class_id)
            students_in_class = selected_class.students.all().order_by('name')
            recorded_attendance = Attendance.objects.filter(school_class=selected_class, date=class_date)
            present_student_ids = list(recorded_attendance.values_list('student_id', flat=True))
            for student in students_in_class:
                attendance_records.append({
                    'student_name': student.name,
                    'status': 'Presente' if student.id in present_student_ids else 'Ausente'
                })
        except SchoolClass.DoesNotExist:
            pass 

    report_title_main = "Relatório de Frequência"
    report_period_str = f"Turma: {selected_class.name if selected_class else 'N/A'} - Data: {DateFormat(class_date).format('d/m/Y')}"

    pdf = PDFReport(church_config=church_config, 
                    report_title=report_title_main, 
                    report_period=report_period_str)
    pdf.add_page()
    current_font_family = pdf.font_family

    line_height = 7
    col_widths_attendance = {"name": 140, "status": 50}

    if attendance_records:
        pdf.set_font(current_font_family, 'B', 9)
        pdf.set_x(pdf.l_margin)
        pdf.cell(col_widths_attendance["name"], line_height, "Nome do Aluno", 1, 0, 'C')
        pdf.cell(col_widths_attendance["status"], line_height, "Status", 1, 1, 'C')

        pdf.set_font(current_font_family, '', 8)
        for record in attendance_records:
            if pdf.get_y() > (pdf.h - pdf.b_margin - line_height * 2):
                pdf.add_page()
                pdf.set_font(current_font_family, 'B', 9)
                pdf.set_x(pdf.l_margin)
                pdf.cell(col_widths_attendance["name"], line_height, "Nome do Aluno", 1, 0, 'C')
                pdf.cell(col_widths_attendance["status"], line_height, "Status", 1, 1, 'C')
                pdf.set_font(current_font_family, '', 8)
            
            start_y = pdf.get_y()
            pdf.set_x(pdf.l_margin)
            pdf.multi_cell(col_widths_attendance["name"], line_height, record['student_name'], 1, 'L')
            max_y = pdf.get_y()
            pdf.set_xy(pdf.l_margin + col_widths_attendance["name"], start_y)
            pdf.multi_cell(col_widths_attendance["status"], line_height, record['status'], 1, 'L')
            max_y = max(max_y, pdf.get_y())
            pdf.set_y(max_y)
    else:
        pdf.set_font(current_font_family, '', 10)
        pdf.set_x(pdf.l_margin)
        pdf.cell(0, line_height, "Nenhum dado de frequência para os filtros selecionados.", 0, 1, 'C')

    pdf.set_font(current_font_family, '', 10)
    pdf_output_bytes = pdf.output(dest='S')
    response = HttpResponse(pdf_output_bytes, content_type='application/pdf')
    response["Content-Disposition"] = f'attachment; filename="frequencia_{selected_class.name.replace(" ", "_")}_{filters["class_date_str"]}.pdf"'

@login_required
def relatorio_financeiro_categorias(request):
    filters = _get_report_filters(request)
    month_param = filters["month_param"]
    year_param = filters["year_param"]

    start_date = date(year_param, month_param, 1)
    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    incomes = Income.objects.filter(date__gte=start_date, date__lte=end_date).select_related("category").order_by("category__name", "date")
    incomes_grouped = {}
    total_revenue = Decimal("0.00")
    for income_item in incomes: # Renamed to avoid conflict
        category_name = income_item.category.name if income_item.category else "Sem Categoria"
        if category_name not in incomes_grouped:
            incomes_grouped[category_name] = {"items": [], "total": Decimal("0.00")}
        incomes_grouped[category_name]["items"].append(income_item)
        incomes_grouped[category_name]["total"] += income_item.amount
        total_revenue += income_item.amount

    expenses = Expense.objects.filter(date__gte=start_date, date__lte=end_date).select_related("category").order_by("category__name", "date")
    expenses_grouped = {}
    total_expenditure = Decimal("0.00")
    for expense_item in expenses: # Renamed to avoid conflict
        category_name = expense_item.category.name if expense_item.category else "Sem Categoria"
        if category_name not in expenses_grouped:
            expenses_grouped[category_name] = {"items": [], "total": Decimal("0.00")}
        expenses_grouped[category_name]["items"].append(expense_item)
        expenses_grouped[category_name]["total"] += expense_item.amount
        total_expenditure += expense_item.amount

    final_balance = total_revenue - total_expenditure

    context = {
        "active_menu": "reports",
        "selected_month": month_param,
        "selected_year": year_param,
        "available_years": filters["available_years"],
        "available_months": filters["available_months"],
        "start_date": start_date,
        "end_date": end_date,
        "incomes_grouped": dict(sorted(incomes_grouped.items())),
        "expenses_grouped": dict(sorted(expenses_grouped.items())),
        "total_revenue": total_revenue,
        "total_expenditure": total_expenditure,
        "final_balance": final_balance,
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }

    return render(request, "reports/financeiro_categorias.html", context)

# --- Relatórios de Membros ---
@login_required
def relatorio_membros_estatisticas(request):
    filters = _get_report_filters(request)
    total_members = Member.objects.count()
    active_members = Member.objects.filter(status="active").count()
    members_by_status = Member.objects.values("status").annotate(count=Count("id")).order_by("-count")
    members_by_gender = Member.objects.values("gender").annotate(count=Count("id")).order_by("-count")
    members_by_marital_status = Member.objects.values("marital_status").annotate(count=Count("id")).order_by("-count")
    members_by_type = Member.objects.values("member_type").annotate(count=Count("id")).order_by("-count")

    status_map = dict(Member.STATUS_CHOICES)
    gender_map = dict(Member.GENDER_CHOICES)
    marital_map = dict(Member.MARITAL_STATUS_CHOICES)
    type_map = dict(Member.MEMBER_TYPE_CHOICES)

    context = {
        "active_menu": "reports",
        "total_members": total_members,
        "active_members": active_members, 
        "members_by_status": [(status_map.get(item["status"], item["status"]), item["count"]) for item in members_by_status],
        "members_by_gender": [(gender_map.get(item["gender"], item["gender"]), item["count"]) for item in members_by_gender],
        "members_by_marital_status": [(marital_map.get(item["marital_status"], item["marital_status"]), item["count"]) for item in members_by_marital_status],
        "members_by_type": [(type_map.get(item["member_type"], item["member_type"]), item["count"]) for item in members_by_type],
        "church_config": filters["church_config"],
    }
    return render(request, "reports/members_estatisticas.html", context)

@login_required
def export_membros_estatisticas_xlsx(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    total_members = Member.objects.count()
    active_members = Member.objects.filter(status="active").count()
    members_by_status = Member.objects.values("status").annotate(count=Count("id")).order_by("-count")
    members_by_gender = Member.objects.values("gender").annotate(count=Count("id")).order_by("-count")
    members_by_marital_status = Member.objects.values("marital_status").annotate(count=Count("id")).order_by("-count")
    members_by_type = Member.objects.values("member_type").annotate(count=Count("id")).order_by("-count")
    status_map = dict(Member.STATUS_CHOICES); gender_map = dict(Member.GENDER_CHOICES)
    marital_map = dict(Member.MARITAL_STATUS_CHOICES); type_map = dict(Member.MEMBER_TYPE_CHOICES)

    wb = Workbook()
    ws = wb.active
    ws.title = "Estatisticas Membros"
    current_row = 1

    if church_config:
        if church_config.logo and hasattr(church_config.logo, "path") and os.path.exists(church_config.logo.path):
            try:
                img = OpenpyxlImage(church_config.logo.path)
                img.height = 75; img.width = 75
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error adding logo to Estatisticas Excel: {e}")
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4) # Increased colspan for longer text
        cell = ws.cell(row=current_row, column=2, value=church_config.church_name or "Nome da Igreja")
        cell.font = Font(bold=True, size=16); cell.alignment = Alignment(horizontal="center")
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
        cell = ws.cell(row=current_row, column=2, value=f"Pastor: {church_config.president_pastor_name or "-"} | Tesoureiro: {church_config.treasurer_name or "-"}")
        cell.font = Font(size=10); cell.alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    title_cell = ws.cell(row=current_row, column=1, value="Relatório de Estatísticas de Membros")
    title_cell.font = Font(bold=True, size=14); title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[current_row].height = 20
    current_row += 2

    ws.append(["Total de Membros:", total_members])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    ws.append(["Membros Ativos:", active_members])
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 2

    def add_stat_section(title, data_list, key_map):
        nonlocal current_row
        ws.cell(row=current_row, column=1, value=title).font = Font(bold=True, size=12)
        current_row += 1
        ws.append(["Item", "Quantidade"])
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2).font = Font(bold=True)
        current_row += 1
        for item_code, count in data_list:
            ws.append([key_map.get(item_code, item_code), count])
            current_row += 1
        current_row += 1 # Spacer

    add_stat_section("Por Status", [(item["status"], item["count"]) for item in members_by_status], status_map)
    add_stat_section("Por Gênero", [(item["gender"], item["count"]) for item in members_by_gender], gender_map)
    add_stat_section("Por Estado Civil", [(item["marital_status"], item["count"]) for item in members_by_marital_status], marital_map)
    add_stat_section("Por Tipo", [(item["member_type"], item["count"]) for item in members_by_type], type_map)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15 # For merged church info
    ws.column_dimensions["D"].width = 15 # For merged church info

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = "attachment; filename=membros_estatisticas.xlsx"
    return response

@login_required
def export_membros_estatisticas_pdf(request):
    filters = _get_report_filters(request)
    church_config = filters["church_config"]
    total_members = Member.objects.count()
    active_members = Member.objects.filter(status="active").count()
    members_by_status = Member.objects.values("status").annotate(count=Count("id")).order_by("-count")
    members_by_gender = Member.objects.values("gender").annotate(count=Count("id")).order_by("-count")
    members_by_marital_status = Member.objects.values("marital_status").annotate(count=Count("id")).order_by("-count")
    members_by_type = Member.objects.values("member_type").annotate(count=Count("id")).order_by("-count")
    status_map = dict(Member.STATUS_CHOICES); gender_map = dict(Member.GENDER_CHOICES)
    marital_map = dict(Member.MARITAL_STATUS_CHOICES); type_map = dict(Member.MEMBER_TYPE_CHOICES)

    context = {
        "total_members": total_members,
        "active_members": active_members,
        "members_by_status": [(status_map.get(item["status"], item["status"]), item["count"]) for item in members_by_status],
        "members_by_gender": [(gender_map.get(item["gender"], item["gender"]), item["count"]) for item in members_by_gender],
        "members_by_marital_status": [(marital_map.get(item["marital_status"], item["marital_status"]), item["count"]) for item in members_by_marital_status],
        "members_by_type": [(type_map.get(item["member_type"], item["member_type"]), item["count"]) for item in members_by_type],
        "church_config": church_config,
        "report_period": "Geral"
    }
    html_string = render_to_string("reports/membros_estatisticas_pdf_template.html", context)
    #html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
    #pdf_file = html.write_pdf()
    response = HttpResponse(None, content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=membros_estatisticas.pdf"
    return response

@login_required
def relatorio_aniversariantes(request):
    filters = _get_report_filters(request)
    month_param = filters["month_param"]
    birthdays = Member.objects.filter(birth_date__month=month_param)\
                                    .annotate(day=ExtractDay("birth_date"))\
                                    .order_by("day", "name")
    month_name = filters["available_months"].get(month_param, "")

    context = {
        "active_menu": "reports",
        "birthdays": birthdays, 
        "selected_month": month_param,
        "month_name": month_name,
        "available_months": filters["available_months"],
        "church_config": filters["church_config"],
    }
    return render(request, "reports/aniversariantes.html", context)

@login_required
def relatorio_contribuicoes_anuais(request):
    filters = _get_report_filters(request)
    year_param = filters["year_param"]
    member_param = filters["member_param"]
    all_members = Member.objects.filter(status="ativo").order_by("name")
    contributions_data = []
    selected_member_name = "Todos"

    if member_param and member_param != "all":
        members_to_query = Member.objects.filter(pk=member_param)
        if members_to_query.exists():
            selected_member_name = members_to_query.first().name
    else:
        members_to_query = all_members # All active members if "all" or no specific member
        member_param = "all" # Ensure it's set for template logic

    for member_obj in members_to_query:
        monthly_contributions = []
        total_annual = Decimal("0.00")
        for month_num in range(1, 13):
            start_of_month = date(year_param, month_num, 1)
            end_of_month = (start_of_month + relativedelta(months=1)) - relativedelta(days=1)
            month_sum = Income.objects.filter(
                member=member_obj,
                date__gte=start_of_month,
                date__lte=end_of_month,
                category__name__iexact="Dízimos"  # Assuming category name for tithe
            ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
            monthly_contributions.append(month_sum)
            total_annual += month_sum
        contributions_data.append({
            "member_name": member_obj.name,
            "monthly_values": monthly_contributions,
            "total_annual": total_annual
        })
    
    total_overall_contribution = sum([item["total_annual"] for item in contributions_data])  

    context = {
        "active_menu": "reports",
        "selected_year": year_param,
        "selected_member_id": member_param if member_param != "all" else "",
        "selected_member_name": selected_member_name,
        "all_members_for_filter": all_members,
        "contributions": contributions_data,
        "available_years": filters["available_years"],
        "months_header": [m[1][:3] for m in filters["available_months"].items()], # Jan, Fev, Mar...
        "church_config": filters["church_config"],
        "total_overall_contribution": total_overall_contribution
    }
    return render(request, "reports/contribuicoes_anuais.html", context)

# Accountability Reports Views (Prestação de Contas)
@method_decorator(login_required, name='dispatch')
class AccountabilityReportListView(ListView):
    model = AccountabilityReport
    template_name = "reports/accountability_list.html"
    context_object_name = "reports"
    paginate_by = 10

    def get_queryset(self):
        return AccountabilityReport.objects.order_by("-year", "-month")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["active_menu"] = "reports"
        context["church_config"] = ChurchConfiguration.objects.first()
        return context

@method_decorator(login_required, name='dispatch')
class AccountabilityReportDetailView(DetailView):
    model = AccountabilityReport
    template_name = "reports/accountability_detail.html"
    context_object_name = "report"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["active_menu"] = "reports"
        context["church_config"] = ChurchConfiguration.objects.first()
        return context

@method_decorator(login_required, name='dispatch')
class AccountabilityReportCreateView(CreateView):
    model = AccountabilityReport
    form_class = AccountabilityReportForm
    template_name = "reports/accountability_form.html"
    success_url = reverse_lazy("reports:accountability_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data["documents_formset"] = AccountabilityDocumentFormSet(self.request.POST, self.request.FILES, prefix="documents")
        else:
            data["documents_formset"] = AccountabilityDocumentFormSet(prefix="documents")
        data["active_menu"] = "reports"
        data["church_config"] = ChurchConfiguration.objects.first()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        documents_formset = context["documents_formset"]
        with transaction.atomic():
            self.object = form.save()
            if documents_formset.is_valid():
                documents_formset.instance = self.object
                documents_formset.save()
            else:
                # If formset is invalid, re-render the form with errors
                # but first, ensure the main form object is not committed if transaction fails
                # However, CreateView typically saves before formset validation here.
                # Consider validating formset first or handling potential rollback.
                messages.error(self.request, "Erro nos documentos anexados.")
                return self.form_invalid(form) # This will re-render with formset errors
        messages.success(self.request, "Relatório de prestação de contas criado com sucesso!")
        return super().form_valid(form)

@method_decorator(login_required, name='dispatch')
class AccountabilityReportUpdateView(UpdateView):
    model = AccountabilityReport
    form_class = AccountabilityReportForm
    template_name = "reports/accountability_form.html"
    success_url = reverse_lazy("reports:accountability_list")

    def get_context_data(self, **kwargs):
        data = super().get_context_data(**kwargs)
        if self.request.POST:
            data["documents_formset"] = AccountabilityDocumentFormSet(self.request.POST, self.request.FILES, instance=self.object, prefix="documents")
        else:
            data["documents_formset"] = AccountabilityDocumentFormSet(instance=self.object, prefix="documents")
        data["active_menu"] = "reports"
        data["church_config"] = ChurchConfiguration.objects.first()
        return data

    def form_valid(self, form):
        context = self.get_context_data()
        documents_formset = context["documents_formset"]
        with transaction.atomic():
            self.object = form.save()
            if documents_formset.is_valid():
                documents_formset.save()
            else:
                messages.error(self.request, "Erro ao atualizar os documentos anexados.")
                return self.form_invalid(form)
        messages.success(self.request, "Relatório atualizado com sucesso!")
        return super().form_valid(form)

@method_decorator(login_required, name='dispatch')
class AccountabilityReportDeleteView(DeleteView):
    model = AccountabilityReport
    template_name = "reports/accountability_confirm_delete.html"
    success_url = reverse_lazy("reports:accountability_list")
    context_object_name = "report"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["active_menu"] = "reports"
        context["church_config"] = ChurchConfiguration.objects.first()
        return context

    def form_valid(self, form):
        messages.success(self.request, "Relatório de prestação de contas excluído com sucesso!")
        return super().form_valid(form)


@login_required
def relatorio_financeiro_categorias(request):
    filters = _get_report_filters(request)
    month_param = filters["month_param"]
    year_param = filters["year_param"]

    start_date = date(year_param, month_param, 1)
    end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    incomes = Income.objects.filter(date__gte=start_date, date__lte=end_date).select_related("category").order_by("category__name", "date")
    incomes_grouped = {}
    total_revenue = Decimal("0.00")
    for income_item in incomes: # Renamed to avoid conflict
        category_name = income_item.category.name if income_item.category else "Sem Categoria"
        if category_name not in incomes_grouped:
            incomes_grouped[category_name] = {"items": [], "total": Decimal("0.00")}
        incomes_grouped[category_name]["items"].append(income_item)
        incomes_grouped[category_name]["total"] += income_item.amount
        total_revenue += income_item.amount

    expenses = Expense.objects.filter(date__gte=start_date, date__lte=end_date).select_related("category").order_by("category__name", "date")
    expenses_grouped = {}
    total_expenditure = Decimal("0.00")
    for expense_item in expenses: # Renamed to avoid conflict
        category_name = expense_item.category.name if expense_item.category else "Sem Categoria"
        if category_name not in expenses_grouped:
            expenses_grouped[category_name] = {"items": [], "total": Decimal("0.00")}
        expenses_grouped[category_name]["items"].append(expense_item)
        expenses_grouped[category_name]["total"] += expense_item.amount
        total_expenditure += expense_item.amount

    final_balance = total_revenue - total_expenditure

    context = {
        "active_menu": "reports",
        "selected_month": month_param,
        "selected_year": year_param,
        "available_years": filters["available_years"],
        "available_months": filters["available_months"],
        "start_date": start_date,
        "end_date": end_date,
        "incomes_grouped": dict(sorted(incomes_grouped.items())),
        "expenses_grouped": dict(sorted(expenses_grouped.items())),
        "total_revenue": total_revenue,
        "total_expenditure": total_expenditure,
        "final_balance": final_balance,
        "filters_query_string": request.GET.urlencode(),
        "church_config": filters["church_config"],
    }

    return render(request, "reports/financeiro_categorias.html", context)


